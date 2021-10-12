import GradientBasedOptimization as gbopt
import openpyxl as pyxl
import time
from QAnsatz import *
from QSubspaceEigensolver import *
import k_nearest_data as k_data
from QMeasure import HadamardTest_Analytical

W_Hamiltonian = Hamiltonian_in_Pauli_String(qubits=3,
                                            unitary=['X0', 'X1', 'X1X0', 'X1Z0', 'Z1X0', 'X2', 'X2X0',
                                                     'X2Z0', 'X2X1', 'X2X1X0', 'X2X1Z0', 'X2Y1Y0', 'X2Z1',
                                                     'X2Z1X0', 'X2Z1Z0', 'Y2Y0', 'Y2Y1', 'Y2Y1Z0', 'Y2Z1Y0',
                                                     'Z2X0', 'Z2X1', 'Z2X1Z0', 'Z2Y1Y0', 'Z2Z1X0'],
                                            coefficient=[6 / 8, 6 / 8, 4 / 8, 2 / 8, -2 / 8, 6 / 8, 2 / 8, -2 / 8,
                                                         2 / 8, 4 / 8, 2 / 8, -4 / 8, -2 / 8, 2 / 8, -2 / 8, 2 / 8,
                                                         -2 / 8, -2 / 8, 2 / 8, -2 / 8, -2 / 8, 2 / 8, -4 / 8, -2 / 8],
                                            hamiltonian_mat=k_data.W)

D_Hamiltonian = Hamiltonian_in_Pauli_String(qubits=3,
                                            unitary=['I2I1I0', 'Z0', 'Z1', 'Z1Z0', 'Z2', 'Z2Z1Z0'],
                                            coefficient=[30 / 8, 2 / 8, -2 / 8, 2 / 8, -4 / 8, 4 / 8, ],
                                            hamiltonian_mat=k_data.D)

I_Hamiltonian = Hamiltonian_in_Pauli_String(qubits=3,
                                            unitary=['I2I1I0'],
                                            coefficient=[1],
                                            hamiltonian_mat=np.eye(2 ** 3))

L_Hamiltonian = D_Hamiltonian - W_Hamiltonian

if __name__ == '__main__':
    eig = 0
    iterations = 5
    state_scale = 3
    layer = 5
    parameter_num = layer * np.math.floor(state_scale / 2) * 12
    timestamp = time.strftime("_%Y%m%d_%H%M%S", time.localtime())
    filename = 'xlsxdata\\data' + timestamp + '.xlsx'
    sigma = 1
    sigma_analytical = [0,
                        0.4495,
                        0.9015,
                        1.0000,
                        1.1949,
                        1.3027,
                        1.5000,
                        1.6514]

    tracers = []
    print('----------------------------------------------')
    print('eig = ' + str(eig))
    init_parameter = [0 for i in range(parameter_num)]
    res_parameter = init_parameter
    tracer_parameter = []
    for it in range(iterations):
        H = L_Hamiltonian - sigma * D_Hamiltonian
        print('eig = ' + str(eig) + '  it = ' + str(it))
        Solver = SubspaceEigSolver_ClassicalEfficientSimulator(Hamiltonian=H,
                                                               ansatze=HardwareEfficientAnsatze_halflayer(state_scale,
                                                                                                          layer,
                                                                                                          res_parameter),
                                                               weight_list=[i + 1 for i in range(eig + 1)])

        if it == 0:
            initvec = np.zeros(2 ** state_scale)
            initvec[0] = 1
            check_circuit = QuantumCircuit(state_scale)
            check_circuit.initialize(initvec, [i for i in range(state_scale)])
            check_circuit.compose(Solver.ansatze.circuit(), [i for i in range(state_scale)], inplace=True)
            job = execute(check_circuit, state_backend)
            result = job.result()
            eigvec = result.get_statevector(check_circuit, decimals=3)
            delta_vec = np.dot(H.hamiltonian_mat, eigvec)
            norm = np.real(np.dot(delta_vec, delta_vec.conj()))
            tracers.append((sigma, norm))

        res_parameter = gbopt.steepest(Solver.getLossFunctionAnalytical,
                                       Solver.GetJacobianAnalytical,
                                       res_parameter,
                                       alpha=0.5,
                                       iters=100,
                                       direct='-',
                                       tol=1e-7)
        tracer_parameter.append(res_parameter)
        ''' get the ith eigenvector '''
        for j in range(eig + 1):
            initvec = np.zeros(2 ** state_scale)
            initvec[j] = 1
            check_circuit = QuantumCircuit(state_scale)
            check_circuit.initialize(initvec, [i for i in range(state_scale)])
            check_circuit.compose(Solver.ansatze.circuit(), [i for i in range(state_scale)], inplace=True)
            job = execute(check_circuit, state_backend)
            result = job.result()
            state = result.get_statevector(check_circuit, decimals=3)
            lamb = np.real(np.dot(np.dot(state.conj(), H.hamiltonian_mat), state))
            print('eig_' + str(j) + ' = ' + str(lamb))

        initvec = np.zeros(2 ** state_scale)
        initvec[0] = 1
        check_circuit = QuantumCircuit(state_scale)
        check_circuit.initialize(initvec, [i for i in range(state_scale)])
        check_circuit.compose(Solver.ansatze.circuit(), [i for i in range(state_scale)], inplace=True)
        ''' update sigma '''
        AE = L_Hamiltonian.ExpectationMeasurement(MeasurementMethod=HadamardTest_Analytical,
                                                  test_circuit=check_circuit,
                                                  active_qubits=[i for i in range(state_scale)])
        BE = D_Hamiltonian.ExpectationMeasurement(MeasurementMethod=HadamardTest_Analytical,
                                                  test_circuit=check_circuit,
                                                  active_qubits=[i for i in range(state_scale)])
        sigma = AE / BE
        job = execute(check_circuit, state_backend)
        result = job.result()
        eigvec = result.get_statevector(check_circuit, decimals=3)
        H_ana = L_Hamiltonian - sigma_analytical[eig] * D_Hamiltonian
        delta_vec = np.dot(H_ana.hamiltonian_mat, eigvec)
        norm = np.real(np.dot(delta_vec, delta_vec.conj()))
        tracers.append((sigma, norm))
        print('sigma = ' + str(sigma))
    print('----------------------------------------------')
    # Create xlsx file
    wb = pyxl.Workbook()
    wb.save(filename)
    wb = pyxl.load_workbook(filename)
    # Record experiment data per experiment
    trace_sheet = wb.create_sheet(title='Experiment Data')
    trace_sheet.cell(1, 1).value = 'timestep'
    trace_sheet.cell(1, 2).value = 'eigval_' + str(eig)
    trace_sheet.cell(1, 3).value = 'norm_' + str(eig)

    for piece in range(len(tracers)):
        trace_sheet.cell(2 + piece, 2).value = tracers[piece][0]
        trace_sheet.cell(2 + piece, 3).value = tracers[piece][1]
    wb.save(filename)
    wb = pyxl.load_workbook(filename)
    sh = wb.create_sheet(title='parameters')
    for i in range(len(tracer_parameter)):
        for j in range(len(tracer_parameter[i])):
            sh.cell(i + 1, 1 + j).value = tracer_parameter[i][j]
    wb.save(filename)

